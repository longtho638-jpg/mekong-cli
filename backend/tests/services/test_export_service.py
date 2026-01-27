import csv
import io
import json
from datetime import datetime

import pytest

from backend.services.export_service import ExportService

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# Mock data
SAMPLE_DATA = [
    {"id": "1", "name": "User A", "email": "a@example.com", "created_at": datetime(2023, 1, 1, 12, 0, 0)},
    {"id": "2", "name": "User B", "email": "b@example.com", "created_at": datetime(2023, 1, 2, 12, 0, 0)},
    {"id": "3", "name": "User C", "email": "c@example.com", "created_at": datetime(2023, 1, 3, 12, 0, 0)},
]

@pytest.fixture
def export_service():
    return ExportService()

class TestExportService:
    def test_export_csv(self, export_service):
        output = export_service.export_data(SAMPLE_DATA, "csv")
        content = output.getvalue().decode("utf-8")

        # Check BOM
        assert content.startswith("\ufeff")

        # Check Headers
        assert "id,name,email,created_at" in content

        # Check Rows
        assert "User A,a@example.com" in content

        # Parse CSV to verify structure
        reader = csv.DictReader(io.StringIO(content.lstrip("\ufeff")))
        rows = list(reader)
        assert len(rows) == 3
        assert rows[0]["name"] == "User A"

    def test_export_json(self, export_service):
        output = export_service.export_data(SAMPLE_DATA, "json")
        content = output.getvalue().decode("utf-8")

        data = json.loads(content)
        assert len(data) == 3
        assert data[0]["name"] == "User A"
        assert "created_at" in data[0]

    def test_export_xlsx(self, export_service):
        if load_workbook is None:
            pytest.skip("openpyxl not installed")

        try:
            output = export_service.export_data(SAMPLE_DATA, "xlsx")

            # Load workbook to verify
            wb = load_workbook(output)
            ws = wb.active

            # Check Headers
            headers = [cell.value for cell in ws[1]]
            assert "id" in headers
            assert "name" in headers

            # Check Data
            assert ws.cell(row=2, column=2).value == "User A"
            assert ws.max_row == 4  # Header + 3 rows

        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_export_empty_data(self, export_service):
        output = export_service.export_data([], "csv")
        content = output.getvalue().decode("utf-8")
        assert content.startswith("\ufeff")
        assert len(content) == 1 # Only BOM

    def test_unsupported_format(self, export_service):
        with pytest.raises(ValueError):
            export_service.export_data(SAMPLE_DATA, "xml")
